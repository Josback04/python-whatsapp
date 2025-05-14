# app/services/excel_service.py

import os
import logging
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
# Importez votre fonction d'upload Drive
try:
    from .drive_service import upload_file_to_drive
except ImportError:
    from drive_service import upload_file_to_drive

# --- Fonctions Helpe pour le style ---
def set_header_style(cell):
    cell.font = Font(bold=True, color="FFFFFF") # Texte blanc
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Fond bleu
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    cell.border = thin_border

def set_data_style(cell, is_currency=False, is_bold=False):
    cell.alignment = Alignment(horizontal='right' if is_currency else 'left', vertical='center')
    if is_currency:
        cell.number_format = '#,##0.00$' # Format monétaire
    if is_bold:
        cell.font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    cell.border = thin_border

# --- Fonction 1: Coûts Variables (CVU) - Mise à jour ---
def generate_costs_excel(user_id, items_list, total_qte_produced):
    """
    Génère Excel CVU basé sur la nouvelle logique, upload et retourne l'URL.
    items_list: [{'name': str, 'global_kg': float, 'global_pt': float}, ...]
    total_qte_produced: float
    """
    excel_path = f"costs_cvu_{user_id}_{os.urandom(4).hex()}.xlsx"
    drive_url = None
    wb = None

    if total_qte_produced == 0: # Éviter division par zéro
        logging.error("total_qte_produced is zero, cannot calculate CVU.")
        return None

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Coût Variable Unitaire"

        # Titre principal
        ws['A1'] = "Calcul du Coût Variable Unitaire (CVU)"
        ws.merge_cells('A1:F1') # Fusionner sur 6 colonnes
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # En-têtes
        headers = ["Élément du CV", "Qté Globale (Kg)", "Prix Total Global ($)", "Prix Unitaire Matière ($/Kg)", "Qté Unitaire (Kg/unité)", "Montant ($/unité)"]
        header_row_num = 3
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row_num, column=col_num, value=header)
            set_header_style(cell)

        # Remplissage des données et calculs
        current_row = header_row_num + 1
        total_cvu = 0.0
        calculated_items = []

        for item in items_list:
            name = item.get('name', 'N/A')
            global_kg = item.get('global_kg', 0.0)
            global_pt = item.get('global_pt', 0.0)

            cost_per_unit_kg = (global_pt / global_kg) if global_kg != 0 else 0
            kg_per_final_unit = global_kg / total_qte_produced
            cost_per_final_unit = global_pt / total_qte_produced # = cost_per_unit_kg * kg_per_final_unit

            calculated_items.append({
                 "name": name,
                 "global_kg": global_kg,
                 "global_pt": global_pt,
                 "cost_per_unit_kg": cost_per_unit_kg,
                 "kg_per_final_unit": kg_per_final_unit,
                 "cost_per_final_unit": cost_per_final_unit
            })
            total_cvu += cost_per_final_unit

            # Écrire dans le fichier Excel
            ws.cell(row=current_row, column=1, value=name)
            ws.cell(row=current_row, column=2, value=global_kg)
            ws.cell(row=current_row, column=3, value=global_pt)
            ws.cell(row=current_row, column=4, value=cost_per_unit_kg)
            ws.cell(row=current_row, column=5, value=kg_per_final_unit)
            ws.cell(row=current_row, column=6, value=cost_per_final_unit)

            # Appliquer le style aux données
            for col_num in range(1, 7):
                 cell = ws.cell(row=current_row, column=col_num)
                 is_currency = col_num in [3, 4, 6]
                 set_data_style(cell, is_currency=is_currency)
                 if col_num in [5]: # Style pour quantité unitaire
                      cell.number_format = '0.00000'

            current_row += 1

        # Ligne Total
        ws.cell(row=current_row, column=1, value="Total Coût Variable Unitaire (CVU)").font = Font(bold=True)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
        total_cell = ws.cell(row=current_row, column=6, value=total_cvu)
        set_data_style(total_cell, is_currency=True, is_bold=True)

        # Ajuster largeur colonnes
        column_widths = {'A': 30, 'B': 18, 'C': 20, 'D': 25, 'E': 25, 'F': 20}
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Sauvegarde et Upload
        wb.save(excel_path)
        logging.info(f"Excel CVU sauvegardé: {excel_path}")
        drive_url = upload_file_to_drive(excel_path)
        # ... (logging succès/échec upload)

    except Exception as e:
        logging.exception(f"Erreur génération Excel CVU pour {user_id}: {e}")
        drive_url = None
    finally:
        # ... (fermeture wb et suppression fichier local) ...
        if wb:
             try: wb.close()
             except Exception: pass
        if os.path.exists(excel_path):
            try: os.remove(excel_path)
            except OSError as e: logging.error(f"Erreur suppression {excel_path}: {e}")

    return drive_url


# --- Fonction 2: Immobilisations - Mise à jour ---
def generate_amortization_excel(user_id, items_list):
    """
    Génère Excel Immobilisations avec quantité, coût total et amortissement, upload et retourne l'URL.
    items_list: [{'name': str, 'unit_cost': float, 'quantity': int, 'lifespan': int}, ...]
    """
    excel_path = f"immobilisations_{user_id}_{os.urandom(4).hex()}.xlsx"
    drive_url = None
    wb = None

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Amortissement Immobilisations"

        ws['A1'] = "Tableau des Immobilisations"
        ws.merge_cells('A1:G1') # 7 colonnes maintenant
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # En-têtes mis à jour
        headers = ["Immobilisation", "Qté", "Prix Unitaire ($)", "Montant Total ($)", "Durée (ans)", "Amort. Annuel ($)", "Amort. Annuel TOTAL ($)"]
        header_row_num = 3
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row_num, column=col_num, value=header)
            set_header_style(cell) # Utiliser la fonction de style

        current_row = header_row_num + 1
        total_annual_amortization_all_items = 0.0
        grand_total_cost = 0.0

        for item in items_list:
            name = item.get('name', 'N/A')
            quantity = item.get('quantity', 0) # Obtenir la quantité
            unit_cost = item.get('cost', 0.0) # Obtenir le coût unitaire
            lifespan = item.get('lifespan', 0)

            # Calculer le coût total et l'amortissement pour cet item
            total_cost_item = unit_cost * quantity
            annual_amortization_item = (total_cost_item / lifespan) if lifespan > 0 else 0

            # Mettre à jour les totaux généraux
            total_annual_amortization_all_items += annual_amortization_item
            grand_total_cost += total_cost_item

            # Écrire les données dans la ligne
            ws.cell(row=current_row, column=1, value=name)
            ws.cell(row=current_row, column=2, value=quantity)          # Colonne Quantité
            ws.cell(row=current_row, column=3, value=unit_cost)         # Colonne Prix Unitaire
            ws.cell(row=current_row, column=4, value=total_cost_item)   # Colonne Montant Total
            ws.cell(row=current_row, column=5, value=lifespan)          # Colonne Durée
            ws.cell(row=current_row, column=6, value=annual_amortization_item) # Colonne Amort. Annuel
            ws.cell(row=current_row, column=7, value=annual_amortization_item) # Répété pour visuel, on pourrait le fusionner/cacher

            # Appliquer le style aux cellules de données
            for col_num in range(1, 8):
                 cell = ws.cell(row=current_row, column=col_num)
                 is_currency = col_num in [3, 4, 6, 7]
                 is_int = col_num in [2, 5]
                 set_data_style(cell, is_currency=is_currency)
                 if is_int: cell.number_format = '0' # Format entier pour Qté et Durée

            current_row += 1

        # Ligne des Totaux
        ws.cell(row=current_row, column=1, value="TOTAUX").font = Font(bold=True)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3) # Fusionner pour le titre Total

        # Afficher le Grand Total Coût
        total_cost_cell = ws.cell(row=current_row, column=4, value=grand_total_cost)
        set_data_style(total_cost_cell, is_currency=True, is_bold=True)

        # Afficher le Total Amortissement Annuel
        total_amort_cell = ws.cell(row=current_row, column=6, value=total_annual_amortization_all_items)
        set_data_style(total_amort_cell, is_currency=True, is_bold=True)
        total_amort_cell_2 = ws.cell(row=current_row, column=7, value=total_annual_amortization_all_items) # Répéter
        set_data_style(total_amort_cell_2, is_currency=True, is_bold=True)


        # Ajuster la largeur des colonnes
        column_widths = {'A': 35, 'B': 10, 'C': 18, 'D': 20, 'E': 12, 'F': 22, 'G': 25}
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Sauvegarde et Upload
        wb.save(excel_path)
        logging.info(f"Excel Amortissement sauvegardé: {excel_path}")
        drive_url = upload_file_to_drive(excel_path)
        if drive_url:
             logging.info(f"Excel Amortissement uploadé: {drive_url}")
        else:
             logging.error(f"Échec upload Excel Amortissement: {excel_path}")


    except Exception as e:
        logging.exception(f"Erreur génération Excel Amortissement pour {user_id}: {e}")
        drive_url = None
    finally:
        # Nettoyage du fichier local
        if wb:
             try: wb.close()
             except Exception: pass # Ignorer les erreurs de fermeture
        if os.path.exists(excel_path):
            try:
                os.remove(excel_path)
                logging.info(f"Fichier local Excel Amortissement supprimé: {excel_path}")
            except OSError as e:
                logging.error(f"Erreur suppression fichier Excel Amortissement {excel_path}: {e}")

    return drive_url

# --- Fonction 3: Coûts Fixes (CFU) - Mise à jour ---
def generate_cfu_excel(user_id, items_list):
    """
    Génère Excel Coûts Fixes annualisés, upload et retourne l'URL.
    items_list: [{'name': str, 'cost_per_period': float, 'period_months': int}, ...]
    """
    excel_path = f"cfu_{user_id}_{os.urandom(4).hex()}.xlsx"
    drive_url = None
    wb = None

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Coûts Fixes Annualisés"

        ws['A1'] = "Tableau des Coûts Fixes Annualisés"
        ws.merge_cells('A1:E1') # 5 colonnes
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        headers = ["Coût Fixe (Nom)", "Coût par Période ($)", "Période (mois)", "Facteur Annuel", "Coût Annuel ($)"]
        header_row_num = 3
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row_num, column=col_num, value=header)
            set_header_style(cell)

        current_row = header_row_num + 1
        total_annual_fixed_cost = 0.0
        calculated_items = []

        for item in items_list:
            name = item.get('name', 'N/A')
            cost_per_period = item.get('cost_per_period', 0.0)
            period_months = item.get('period_months', 0)

            annual_factor = (12 / period_months) if period_months > 0 else 0
            annual_cost = cost_per_period * annual_factor
            total_annual_fixed_cost += annual_cost

            calculated_items.append({
                "name": name,
                "cost_per_period": cost_per_period,
                "period_months": period_months,
                "annual_factor": annual_factor,
                "annual_cost": annual_cost
            })

            ws.cell(row=current_row, column=1, value=name)
            ws.cell(row=current_row, column=2, value=cost_per_period)
            ws.cell(row=current_row, column=3, value=period_months)
            ws.cell(row=current_row, column=4, value=annual_factor)
            ws.cell(row=current_row, column=5, value=annual_cost)

            # Style
            for col_num in range(1, 6):
                 cell = ws.cell(row=current_row, column=col_num)
                 is_currency = col_num in [2, 5]
                 is_int = col_num == 3
                 is_factor = col_num == 4
                 set_data_style(cell, is_currency=is_currency)
                 if is_int: cell.number_format = '0'
                 if is_factor: cell.number_format = '0.00'


            current_row += 1

        # Ligne Total
        ws.cell(row=current_row, column=1, value="Total Coûts Fixes Annuels").font = Font(bold=True)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        total_cell = ws.cell(row=current_row, column=5, value=total_annual_fixed_cost)
        set_data_style(total_cell, is_currency=True, is_bold=True)

        # Ajuster largeur
        column_widths = {'A': 35, 'B': 20, 'C': 15, 'D': 18, 'E': 20}
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Sauvegarde et Upload
        wb.save(excel_path)
        logging.info(f"Excel CFU sauvegardé: {excel_path}")
        drive_url = upload_file_to_drive(excel_path)
        # ... (logging succès/échec upload)

    except Exception as e:
        logging.exception(f"Erreur génération Excel CFU pour {user_id}: {e}")
        drive_url = None
    finally:
        # ... (fermeture wb et suppression fichier local) ...
        if wb:
             try: wb.close()
             except Exception: pass
        if os.path.exists(excel_path):
            try: os.remove(excel_path)
            except OSError as e: logging.error(f"Erreur suppression {excel_path}: {e}")

    return drive_url

# Note: La fonction generate_costs_env_excel n'est pas mise à jour car son objectif
# et la structure de ses données d'entrée ('env') n'étaient pas clairs.
# Si vous en avez besoin, veuillez clarifier ce qu'elle doit calculer et afficher.